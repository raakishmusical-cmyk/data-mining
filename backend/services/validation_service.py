import os
import re
import uuid
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from backend.repositories.file_repository import file_repo
from backend.repositories.history_repository import history_repo
from backend.services.history_service import history_service
from backend.utils.logger import logger
from backend.utils.normalizer import clean_text, normalize_string, clean_phone_digits, extract_domain, is_valid_val, normalize_phone, normalize_email, normalize_website
from backend.utils.phone import parse_and_validate_phone
from backend.utils.email import is_valid_email
from backend.utils.industry import determine_industry
from backend.utils.tags import determine_tag

HEADERS = [
    "Organization Name", "Salutation", "First Name", "Last Name", "Title",
    "Email", "Secondary Email", "Phone", "Mobile", "Fax", "Skype ID",
    "Website", "Instagram", "Facebook", "LinkedIn", "Twitter", "YouTube",
    "Street", "City", "State", "Zip Code", "Country", "Industry"
]

class DSU:
    def __init__(self, n):
        self.parent = list(range(n))
        
    def find(self, i):
        path = []
        while self.parent[i] != i:
            path.append(i)
            i = self.parent[i]
        for node in path:
            self.parent[node] = i
        return i
        
    def union(self, i, j):
        root_i = self.find(i)
        root_j = self.find(j)
        if root_i != root_j:
            self.parent[root_i] = root_j
            return True
        return False

PRIORITY_MAP = {
    "Phone Duplicate": 1,
    "Mobile Duplicate": 2,
    "Email Duplicate": 3,
    "Website Duplicate": 4,
    "Google Place ID Duplicate": 5,
    "Name + Address Duplicate": 6,
    "Name + Phone Duplicate": 7,
    "Name + Email Duplicate": 8,
    "Name + Website Duplicate": 9
}

# Normalization helpers are imported from backend.utils.normalizer


def normalize_place_id(place_id_val) -> str:
    if not place_id_val:
        return ""
    val = str(place_id_val).strip()
    if val.lower() in ("n/a", "na", "nan", ""):
        return ""
    return val

def get_row_metadata(row, idx, df_columns) -> dict:
    # 1. Completed fields count
    completed_count = 0
    for col in df_columns:
        val = str(row.get(col, "")).strip()
        if val and val.lower() not in ("n/a", "na", "nan", ""):
            completed_count += 1
            
    # 2. Validation Score
    val_score = None
    for col_name in ["Validation Score", "validation_score", "val_score"]:
        if col_name in row:
            try:
                val_score = float(row[col_name])
                break
            except Exception:
                pass
    if val_score is None:
        # Calculate dynamic validation score if not present
        val_score = 0.0
        email = str(row.get("Email", ""))
        if is_valid_email(email):
            val_score += 3.0
        phone = str(row.get("Phone", ""))
        mobile = str(row.get("Mobile", ""))
        if normalize_phone(phone):
            val_score += 2.0
        if normalize_phone(mobile):
            val_score += 2.0
            
    # 3. Quality Score
    qual_score = None
    for col_name in ["Quality Score", "quality_score", "qual_score"]:
        if col_name in row:
            try:
                qual_score = float(row[col_name])
                break
            except Exception:
                pass
    if qual_score is None:
        # Calculate dynamic quality score if not present
        qual_score = 0.0
        web = str(row.get("Website", ""))
        if normalize_website(web):
            qual_score += 2.0
        qual_score += completed_count * 0.5
        
    return {
        "completed_count": completed_count,
        "validation_score": val_score,
        "quality_score": qual_score,
        "row_index": idx
    }

def get_direct_match_details(r1_norm, r2_norm):
    if is_valid_val(r1_norm["phone"]) and r1_norm["phone"] == r2_norm["phone"]:
        return "Phone Matched", ["Phone"]
    if is_valid_val(r1_norm["mobile"]) and r1_norm["mobile"] == r2_norm["mobile"]:
        return "Mobile Matched", ["Mobile"]
    if is_valid_val(r1_norm["email"]) and r1_norm["email"] == r2_norm["email"]:
        return "Email Matched", ["Email"]
    if is_valid_val(r1_norm["website"]) and r1_norm["website"] == r2_norm["website"]:
        return "Website Matched", ["Website"]
    if is_valid_val(r1_norm["facebook"]) and r1_norm["facebook"] == r2_norm["facebook"]:
        return "Facebook Matched", ["Facebook"]
    if is_valid_val(r1_norm["instagram"]) and r1_norm["instagram"] == r2_norm["instagram"]:
        return "Instagram Matched", ["Instagram"]
    if is_valid_val(r1_norm["linkedin"]) and r1_norm["linkedin"] == r2_norm["linkedin"]:
        return "LinkedIn Matched", ["LinkedIn"]
    if is_valid_val(r1_norm["twitter"]) and r1_norm["twitter"] == r2_norm["twitter"]:
        return "Twitter Matched", ["Twitter"]
    if is_valid_val(r1_norm["youtube"]) and r1_norm["youtube"] == r2_norm["youtube"]:
        return "YouTube Matched", ["YouTube"]
    return None, []

def get_match_reason(r1, r2) -> str:
    from backend.utils.normalizer import clean_phone_digits, normalize_string, extract_domain
    
    p1 = clean_phone_digits(r1.get("Phone", ""))
    p2 = clean_phone_digits(r2.get("Phone", ""))
    m1 = clean_phone_digits(r1.get("Mobile", ""))
    m2 = clean_phone_digits(r2.get("Mobile", ""))
    e1 = str(r1.get("Email", "")).strip().lower()
    e2 = str(r2.get("Email", "")).strip().lower()
    w1 = extract_domain(r1.get("Website", ""))
    w2 = extract_domain(r2.get("Website", ""))
    pl1 = str(r1.get("Google Place ID", "")).strip()
    pl2 = str(r2.get("Google Place ID", "")).strip()
    n1 = normalize_string(r1.get("Organization Name", ""))
    n2 = normalize_string(r2.get("Organization Name", ""))
    a1 = normalize_string(r1.get("Street", ""))
    a2 = normalize_string(r2.get("Street", ""))
    
    reasons = []
    if is_valid_val(p1) and p1 == p2:
        reasons.append("Phone Duplicate")
    if is_valid_val(m1) and m1 == m2:
        reasons.append("Mobile Duplicate")
    if is_valid_val(e1) and e1 == e2:
        reasons.append("Email Duplicate")
    if is_valid_val(w1) and w1 == w2:
        reasons.append("Website Duplicate")
    if is_valid_val(pl1) and pl1 == pl2:
        reasons.append("Google Place ID Duplicate")
    if is_valid_val(n1) and n1 == n2:
        if is_valid_val(a1) and a1 == a2:
            reasons.append("Name + Address Duplicate")
        if is_valid_val(p1) and p1 == p2:
            reasons.append("Name + Phone Duplicate")
        if is_valid_val(e1) and e1 == e2:
            reasons.append("Name + Email Duplicate")
        if is_valid_val(w1) and w1 == w2:
            reasons.append("Name + Website Duplicate")
            
    if not reasons:
        return None
        
    return min(reasons, key=lambda x: PRIORITY_MAP[x])

def clean_for_fuzzy(text: str) -> str:
    if not text:
        return ""
    # Lowercase
    text = str(text).lower().strip()
    # Remove punctuation
    text = re.sub(r"[^\w\s]", "", text)
    # Normalize multiple spaces to single space
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def check_name_keyword_match(name: str, keyword: str) -> bool:
    from rapidfuzz import fuzz
    if not name or not keyword:
        return False
    
    c_name = clean_for_fuzzy(name)
    c_kw = clean_for_fuzzy(keyword)
    
    if not c_name or not c_kw:
        return False
        
    # 1. Check if clean keyword is a substring of clean name
    if c_kw in c_name:
        return True
        
    # 2. Check if clean name is a substring of clean keyword
    if c_name in c_kw:
        return True
        
    # 3. Check rapidfuzz token_set_ratio
    score = fuzz.token_set_ratio(c_name, c_kw)
    if score >= 75.0:
        return True
        
    # 4. Check if any word of the keyword is in the name (excluding generic stop words)
    stop_words = {"shop", "store", "academy", "center", "club", "studio", "station", "office", "department", "company", "group", "ltd", "pvt", "limited", "and", "the", "of", "in", "at", "on", "for"}
    kw_words = [w for w in c_kw.split() if w not in stop_words and len(w) > 2]
    name_words = set(c_name.split())
    for w in kw_words:
        if w in name_words:
            return True
            
    return False

class ValidationService:
    def __init__(self):
        # Maps validation session ID -> dict of original/working file paths
        self.sessions = {}

    async def create_session(self, filename: str, content: bytes) -> str:
        """Stores the uploaded validation file on disk and registers validation session."""
        session_id = str(uuid.uuid4())
        
        # Determine upload extension
        ext = os.path.splitext(filename)[1].lower()
        if ext not in (".csv", ".xlsx", ".xls"):
            raise ValueError("Unsupported file format. Please upload CSV, XLSX, or XLS.")
            
        # Target path inside Output/Validation
        file_repo.ensure_directories()
        unique_filename = f"{session_id}{ext}"
        filepath = os.path.join("Output/Validation", unique_filename)
        
        # Save file to disk
        with open(filepath, "wb") as f:
            f.write(content)
            
        # Extract basic info
        df = self._read_file(filepath)
        row_count = len(df)
        
        # Register in File History
        db_file_id = await history_service.register_file(
            filename=filename,
            category="Mining Files", # initially labeled as raw files
            district="Validation Upload",
            keyword_count=0,
            row_count=row_count,
            filepath=filepath
        )
        
        # Save session link
        self.sessions[session_id] = {
            "file_id": db_file_id,
            "original_name": filename,
            "original_path": filepath,
            "current_working_path": filepath,
            "ext": ext,
            "row_count": row_count
        }
        
        logger.info(f"Created validation session {session_id} for file {filename}")
        return session_id

    def _read_file(self, filepath: str) -> pd.DataFrame:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".csv":
            return pd.read_csv(filepath).fillna("N/A")
        else:
            return pd.read_excel(filepath).fillna("N/A")

    async def scan_duplicates(self, session_id: str) -> dict:
        """
        Scans current working file for duplicates using a priority-based matching engine on contact details.
        Clusters matching rows and populates duplicate summaries and preview tables.
        """
        session = self.sessions.get(session_id)
        if not session:
            raise ValueError("Session not found.")
            
        df = self._read_file(session["current_working_path"])
        n = len(df)
        dsu = DSU(n)
        
        # Extract and normalize fields for all rows
        normalized_data = []
        for idx, row in df.iterrows():
            normalized_data.append({
                "index": idx,
                "row_number": idx + 2,
                "phone": normalize_phone(row.get("Phone", "")),
                "mobile": normalize_phone(row.get("Mobile", "")),
                "email": normalize_email(row.get("Email", "")),
                "website": normalize_website(row.get("Website", "")),
                "facebook": normalize_website(row.get("Facebook", "")),
                "instagram": normalize_website(row.get("Instagram", "")),
                "linkedin": normalize_website(row.get("LinkedIn", "")),
                "twitter": normalize_website(row.get("Twitter", "")),
                "youtube": normalize_website(row.get("YouTube", "")),
                "name": normalize_string(row.get("Organization Name", "")),
                "address": normalize_string(row.get("Street", ""))
            })
            
        # Lookup maps for matching
        phone_map = {}
        mobile_map = {}
        email_map = {}
        website_map = {}
        facebook_map = {}
        instagram_map = {}
        linkedin_map = {}
        twitter_map = {}
        youtube_map = {}
        
        # parent_match stores for each duplicate row: (matched_idx, reason, matched_fields)
        parent_match = {}
        
        for idx in range(n):
            r = normalized_data[idx]
            
            is_phone_valid = is_valid_val(r["phone"])
            is_mobile_valid = is_valid_val(r["mobile"])
            is_email_valid = is_valid_val(r["email"])
            is_website_valid = is_valid_val(r["website"])
            is_facebook_valid = is_valid_val(r["facebook"])
            is_instagram_valid = is_valid_val(r["instagram"])
            is_linkedin_valid = is_valid_val(r["linkedin"])
            is_twitter_valid = is_valid_val(r["twitter"])
            is_youtube_valid = is_valid_val(r["youtube"])
            
            match_found = False
            matched_idx = None
            reason = None
            matched_fields = []
            
            # Check matches in priority order of identifiers
            if not match_found and is_phone_valid and r["phone"] in phone_map:
                matched_idx = phone_map[r["phone"]][0]
                reason = "Phone Matched"
                matched_fields = ["Phone"]
                match_found = True
                
            if not match_found and is_mobile_valid and r["mobile"] in mobile_map:
                matched_idx = mobile_map[r["mobile"]][0]
                reason = "Mobile Matched"
                matched_fields = ["Mobile"]
                match_found = True
                
            if not match_found and is_email_valid and r["email"] in email_map:
                matched_idx = email_map[r["email"]][0]
                reason = "Email Matched"
                matched_fields = ["Email"]
                match_found = True
                
            if not match_found and is_website_valid and r["website"] in website_map:
                matched_idx = website_map[r["website"]][0]
                reason = "Website Matched"
                matched_fields = ["Website"]
                match_found = True
                
            if not match_found and is_facebook_valid and r["facebook"] in facebook_map:
                matched_idx = facebook_map[r["facebook"]][0]
                reason = "Facebook Matched"
                matched_fields = ["Facebook"]
                match_found = True
                
            if not match_found and is_instagram_valid and r["instagram"] in instagram_map:
                matched_idx = instagram_map[r["instagram"]][0]
                reason = "Instagram Matched"
                matched_fields = ["Instagram"]
                match_found = True
                
            if not match_found and is_linkedin_valid and r["linkedin"] in linkedin_map:
                matched_idx = linkedin_map[r["linkedin"]][0]
                reason = "LinkedIn Matched"
                matched_fields = ["LinkedIn"]
                match_found = True
                
            if not match_found and is_twitter_valid and r["twitter"] in twitter_map:
                matched_idx = twitter_map[r["twitter"]][0]
                reason = "Twitter Matched"
                matched_fields = ["Twitter"]
                match_found = True
                
            if not match_found and is_youtube_valid and r["youtube"] in youtube_map:
                matched_idx = youtube_map[r["youtube"]][0]
                reason = "YouTube Matched"
                matched_fields = ["YouTube"]
                match_found = True
                
            if match_found:
                dsu.union(idx, matched_idx)
                parent_match[idx] = (matched_idx, reason, matched_fields)
                
            # Register in maps
            if is_phone_valid:
                phone_map.setdefault(r["phone"], []).append(idx)
            if is_mobile_valid:
                mobile_map.setdefault(r["mobile"], []).append(idx)
            if is_email_valid:
                email_map.setdefault(r["email"], []).append(idx)
            if is_website_valid:
                website_map.setdefault(r["website"], []).append(idx)
            if is_facebook_valid:
                facebook_map.setdefault(r["facebook"], []).append(idx)
            if is_instagram_valid:
                instagram_map.setdefault(r["instagram"], []).append(idx)
            if is_linkedin_valid:
                linkedin_map.setdefault(r["linkedin"], []).append(idx)
            if is_twitter_valid:
                twitter_map.setdefault(r["twitter"], []).append(idx)
            if is_youtube_valid:
                youtube_map.setdefault(r["youtube"], []).append(idx)
                
        # Group by DSU parent
        groups = {}
        for idx in range(n):
            root = dsu.find(idx)
            groups.setdefault(root, []).append(idx)
            
        keep_indices = set()
        duplicate_indices = set()
        row_to_keep_record = {} # duplicate_row_idx -> keep_row_idx
        
        for root, indices in groups.items():
            if len(indices) == 1:
                keep_indices.add(indices[0])
            else:
                row_metadatas = [get_row_metadata(df.iloc[idx], idx, df.columns) for idx in indices]
                row_metadatas.sort(key=lambda item: (
                    -item["completed_count"],
                    -item["validation_score"],
                    -item["quality_score"],
                    item["row_index"]
                ))
                best_idx = row_metadatas[0]["row_index"]
                keep_indices.add(best_idx)
                for idx in indices:
                    if idx != best_idx:
                        duplicate_indices.add(idx)
                        row_to_keep_record[idx] = best_idx
                        
        # Construct output preview table rows
        rows_preview = []
        for idx, row in df.iterrows():
            is_dup = idx in duplicate_indices
            
            if is_dup:
                keep_idx = row_to_keep_record[idx]
                direct_reason, direct_fields = get_direct_match_details(normalized_data[idx], normalized_data[keep_idx])
                if direct_reason:
                    reason = direct_reason
                    matched_fields = direct_fields
                else:
                    parent_idx, parent_reason, parent_fields = parent_match.get(idx, (keep_idx, "Duplicate", []))
                    reason = parent_reason
                    matched_fields = parent_fields
                matched_row_num = keep_idx + 2
                g_id = keep_idx + 2
            else:
                reason = "Original Keep"
                matched_row_num = None
                matched_fields = []
                g_id = idx + 2
                
            meta = get_row_metadata(row, idx, df.columns)
            val_score = meta["validation_score"]
            qual_score = meta["quality_score"]
            
            row_preview_dict = {
                "row_number": idx + 2,
                "Row Number": idx + 2,
                "group_id": g_id,
                "Group ID": g_id,
                "keep": not is_dup,
                "Keep": not is_dup,
                "duplicate_reason": reason,
                "Duplicate Reason": reason,
                "matched_row_number": matched_row_num,
                "Matched Row Number": matched_row_num,
                "matched_fields": matched_fields,
                "Matched Fields": matched_fields,
                "validation_score": val_score,
                "Validation Score": val_score,
                "quality_score": qual_score,
                "Quality Score": qual_score,
                "name": str(row.get("Organization Name", "N/A")),
                "salutation": str(row.get("Salutation", "N/A")),
                "first_name": str(row.get("First Name", "N/A")),
                "last_name": str(row.get("Last Name", "N/A")),
                "title": str(row.get("Title", "N/A")),
                "email": str(row.get("Email", "N/A")),
                "secondary_email": str(row.get("Secondary Email", "N/A")),
                "phone": str(row.get("Phone", "N/A")),
                "mobile": str(row.get("Mobile", "N/A")),
                "fax": str(row.get("Fax", "N/A")),
                "skype_id": str(row.get("Skype ID", "N/A")),
                "website": str(row.get("Website", "N/A")),
                "instagram": str(row.get("Instagram", "N/A")),
                "facebook": str(row.get("Facebook", "N/A")),
                "linkedin": str(row.get("LinkedIn", "N/A")),
                "twitter": str(row.get("Twitter", "N/A")),
                "youtube": str(row.get("YouTube", "N/A")),
                "street": str(row.get("Street", "N/A")),
                "city": str(row.get("City", "N/A")),
                "state": str(row.get("State", "N/A")),
                "zip_code": str(row.get("Zip Code", "N/A")),
                "country": str(row.get("Country", "N/A")),
                "industry": str(row.get("Industry", "N/A")),
                "keyword": str(row.get("Keyword", "N/A")),
                "tags": str(row.get("Tags", "N/A")),
            }
            rows_preview.append(row_preview_dict)
            
        # Summary counts - calculated independently
        total_rows = len(df)
        dup_groups = len([g for g, ind in groups.items() if len(ind) > 1])
        dup_rows = len(duplicate_indices)
        
        phone_duplicates_count = 0
        mobile_duplicates_count = 0
        email_duplicates_count = 0
        website_duplicates_count = 0
        business_name_duplicates_count = 0
        address_duplicates_count = 0
        
        for idx in duplicate_indices:
            keep_idx = row_to_keep_record[idx]
            r_dup = normalized_data[idx]
            r_keep = normalized_data[keep_idx]
            
            if is_valid_val(r_dup["phone"]) and r_dup["phone"] == r_keep["phone"]:
                phone_duplicates_count += 1
            if is_valid_val(r_dup["mobile"]) and r_dup["mobile"] == r_keep["mobile"]:
                mobile_duplicates_count += 1
            if is_valid_val(r_dup["email"]) and r_dup["email"] == r_keep["email"]:
                email_duplicates_count += 1
            if is_valid_val(r_dup["website"]) and r_dup["website"] == r_keep["website"]:
                website_duplicates_count += 1
            if is_valid_val(r_dup["name"]) and r_dup["name"] == r_keep["name"]:
                business_name_duplicates_count += 1
            if is_valid_val(r_dup["address"]) and r_dup["address"] == r_keep["address"]:
                address_duplicates_count += 1
                
        return {
            "summary": {
                "rows_scanned": total_rows,
                "Rows Scanned": total_rows,
                "duplicate_groups": dup_groups,
                "Duplicate Groups": dup_groups,
                "duplicate_rows": dup_rows,
                "Duplicate Rows": dup_rows,
                "phone_duplicates": phone_duplicates_count,
                "Phone Duplicates": phone_duplicates_count,
                "mobile_duplicates": mobile_duplicates_count,
                "Mobile Duplicates": mobile_duplicates_count,
                "email_duplicates": email_duplicates_count,
                "Email Duplicates": email_duplicates_count,
                "website_duplicates": website_duplicates_count,
                "Website Duplicates": website_duplicates_count,
                "business_name_duplicates": business_name_duplicates_count,
                "Business Name Duplicates": business_name_duplicates_count,
                "address_duplicates": address_duplicates_count,
                "Address Duplicates": address_duplicates_count
            },
            "preview": rows_preview
        }

    def _determine_best_row(self, df: pd.DataFrame, indices: list[int]) -> int:
        """Determines best quality row from list of indices."""
        best_idx = indices[0]
        max_score = -1
        
        for idx in indices:
            row = df.iloc[idx]
            score = 0
            
            # 1. Filled columns check
            for col in df.columns:
                val = str(row[col]).strip()
                if val and val.lower() not in ("n/a", "na", ""):
                    score += 1
                    
            # 2. Valid email check
            email = str(row.get("Email", ""))
            if is_valid_email(email):
                score += 3
                
            # 3. Valid website check
            web = str(row.get("Website", ""))
            if web and web.lower() not in ("n/a", "na", ""):
                score += 2
                
            # 4. Valid phone/mobile check
            phone = str(row.get("Phone", ""))
            mobile = str(row.get("Mobile", ""))
            if phone != "N/A" or mobile != "N/A":
                score += 2
                
            if score > max_score:
                max_score = score
                best_idx = idx
                
        return best_idx

    async def mark_duplicates(self, session_id: str) -> str:
        """Highlights duplicate rows in Yellow, Bold. Generates marked.xlsx."""
        session = self.sessions.get(session_id)
        if not session:
            raise ValueError("Session not found.")
            
        df = self._read_file(session["current_working_path"])
        
        # Re-run scan to identify duplicate rows indices
        res = await self.scan_duplicates(session_id)
        duplicate_excel_rows = [r["row_number"] for r in res["preview"] if not r["keep"]]
        
        # Create Excel
        orig_name = session["original_name"]
        marked_name = f"marked_{orig_name.replace('.csv', '.xlsx')}"
        if not marked_name.endswith(".xlsx"):
            marked_name += ".xlsx"
            
        marked_path = os.path.join("Output/Marked", marked_name)
        file_repo.ensure_directories()
        
        # Write clean dataframe to temporary excel first
        for col in HEADERS:
            if col not in df.columns:
                df[col] = "N/A"
        df = df[HEADERS]
        df.to_excel(marked_path, index=False)
        
        # Load workbook and apply styles
        wb = load_workbook(marked_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        bold_font = Font(bold=True)
        
        for r_num in duplicate_excel_rows:
            # Highlight all cells in the row
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r_num, column=col)
                cell.fill = yellow_fill
                cell.font = bold_font
                
        wb.save(marked_path)
        
        # Register in History
        file_id = await history_service.register_file(
            filename=marked_name,
            category="Marked Files",
            district="Validation Mark",
            keyword_count=0,
            row_count=len(df),
            filepath=marked_path
        )
        
        return file_id

    async def delete_duplicates(self, session_id: str, delete_row_numbers: list[int] = None) -> str:
        """Deletes duplicate rows. Keeps best records or deletes selected. Generates clean.xlsx."""
        session = self.sessions.get(session_id)
        if not session:
            raise ValueError("Session not found.")
            
        df = self._read_file(session["current_working_path"])
        
        if delete_row_numbers is not None:
            # delete_row_numbers are 1-indexed Excel row numbers (header is 1, data starts at 2).
            # Convert them to 0-based DataFrame indices.
            indices_to_delete = {int(r_num) - 2 for r_num in delete_row_numbers}
            keep_indices = [idx for idx in range(len(df)) if idx not in indices_to_delete]
        else:
            # Re-run scan to identify keep rows indices
            res = await self.scan_duplicates(session_id)
            keep_indices = [r["row_number"] - 2 for r in res["preview"] if r["keep"]]
        
        # Slice dataframe
        clean_df = df.iloc[keep_indices]
        for col in HEADERS:
            if col not in clean_df.columns:
                clean_df[col] = "N/A"
        clean_df = clean_df[HEADERS]
        
        orig_name = session["original_name"]
        clean_name = f"clean_{orig_name.replace('.csv', '.xlsx')}"
        if not clean_name.endswith(".xlsx"):
            clean_name += ".xlsx"
            
        clean_path = os.path.join("Output/Cleaned", clean_name)
        file_repo.ensure_directories()
        
        # Save Excel
        clean_df.to_excel(clean_path, index=False)
        
        # Auto size columns
        wb = load_workbook(clean_path)
        ws = wb.active
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        wb.save(clean_path)
        
        # Register in Recent Files
        file_id = await history_service.register_file(
            filename=clean_name,
            category="Cleaned Files",
            district="Validation Cleaned",
            keyword_count=0,
            row_count=len(clean_df),
            filepath=clean_path
        )
        
        # Update working path to clean.xlsx
        session["current_working_path"] = clean_path
        
        return file_id

    async def fix_industry_and_tags(self, session_id: str) -> str:
        """Updates standard industries and tags in-place on current working file."""
        session = self.sessions.get(session_id)
        if not session:
            raise ValueError("Session not found.")
            
        working_path = session["current_working_path"]
        df = self._read_file(working_path)
        
        # Determine the fallback/original keyword from the spreadsheet
        original_keyword = "Sports Lead"
        if "Keyword" in df.columns:
            kw_counts = df["Keyword"].dropna().astype(str).tolist()
            valid_kws = [k.strip() for k in kw_counts if k.strip().lower() not in ("", "n/a", "na", "nan")]
            if valid_kws:
                from collections import Counter
                original_keyword = Counter(valid_kws).most_common(1)[0][0]
        
        # Iterate and fix in-place
        for idx, row in df.iterrows():
            # Get keyword for this row, fallback to original_keyword
            kw = str(row.get("Keyword", "")).strip()
            if kw.lower() in ("", "n/a", "na", "nan"):
                kw = original_keyword
                
            df.at[idx, "Industry"] = kw
            df.at[idx, "Keyword"] = kw
            df.at[idx, "Tags"] = "Sports Lead"
                
        # Save back to same path
        for col in HEADERS:
            if col not in df.columns:
                df[col] = "N/A"
        df = df[HEADERS]
        
        ext = os.path.splitext(working_path)[1].lower()
        if ext == ".csv":
            df.to_csv(working_path, index=False)
        else:
            df.to_excel(working_path, index=False)
            
        # Re-register processed file update in History database
        processed_name = f"processed_{os.path.basename(working_path)}"
        processed_path = os.path.join("Output/Cleaned", processed_name)
        
        # Copy to processed path
        if ext == ".csv":
            df.to_csv(processed_path, index=False)
        else:
            df.to_excel(processed_path, index=False)
            
        file_id = await history_service.register_file(
            filename=processed_name,
            category="Processed Files",
            district="Classification Fix",
            keyword_count=0,
            row_count=len(df),
            filepath=processed_path
        )
        
        session["current_working_path"] = processed_path
        
        return file_id

validation_service = ValidationService()
